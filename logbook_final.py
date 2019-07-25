from tkinter import *
from tkinter import ttk
import pandas as pd
import tkinter as tk
import xlsxwriter
from xlrd import open_workbook
import xlwt
import os.path
import os
from datetime import datetime
import numpy as np
import xlwings
from openpyxl import workbook
from openpyxl import load_workbook
import win32com.client as win32
from fpdf import FPDF




fname=0
tname=0
avalu=0
bvalu=0
cvalu=0
dvalu=0
evalu=0
fvalu=0
gvalu=0
hvalu=0
ivalu=0
jvalu=0
kvalu=0
lvalu=0
mvalu=0
nvalu=0
ovalu=0
pvalu=0
qvalu=0
rvalu=0
svalu=0
tvalu=0
uvalu=0
vvalu=None
wvalu=0
tmp=0
ind=0
indind=0


now = datetime.now()
window=Tk()
window.title("Document Entry")
window.geometry("1300x700")
window.configure(background="white")


def pdf():
    global ind,tname,fname,pdfvar,excelvar,avalu,bvalu,cvalu,dvalu,evalu,fvalu,gvalu,hvalu,ivalu,jvalu,kvalu,lvalu,mvalu,nvalu,ovalu,pvalu,qvalu,rvalu,svalu,tvalu,uvalu,vvalu,wvalu
    if excelvar.get()!="Do You Want To Save Excel File?":
        if len(tvalu)==0:
            dataframe=pd.read_excel(tname)
            dataframe=dataframe['S.No.'].tolist()
            if len(dataframe)==0:
                ind=1
            else:
                dataframe=[int(x) for x in dataframe if not np.isnan(x)]
                dataframe=max(dataframe)
                ind=dataframe
        else:
            ind=tvalu
        if ivalu==None:
            ivalu=' '
        if jvalu==None:
            jvalu=' '
        if vvalu==None:
            vvalu=' '
        pdf = FPDF()
        pdf.set_auto_page_break(True,margin=0.5)
        pdf.add_page()
        pdf.set_font("Arial", size=26,style="B")
        pdf.cell(180, 30, txt="SUMMARY!!!!", ln=1, align="C")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="S.NO.", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(ind), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="First Party Name", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(avalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="First Party Mobile Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(bvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="First Party Aadhar Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(cvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="First Party Pan Card", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(dvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Second Party Name", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(evalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Second Party Mobile Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(fvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Second Party Aadhar Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(gvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Second Party Pan Card", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(hvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Document", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(ivalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Sub Registrar", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(jvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Witness 1 Name", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(kvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Witness 1 Mobile Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(lvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Witness 2 Name", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(mvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Witness 2 Mobile Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(nvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Document Date", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(ovalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Document Day", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(vvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Deed Writer Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(pvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Registration Number", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(qvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Property Value", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(rvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Maarfat", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(svalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Entry Time", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="B")
        pdf.cell(60, 8, txt=str(wvalu), ln=1, align="L")
        pdf.set_font("Arial", size=16)
        pdf.cell(90, 8, txt="Property Address", ln=0, align="L")
        pdf.cell(12, 8, txt=":", ln=0, align="L")
        pdf.set_font("Arial",style="BI")
        pdf.multi_cell(90, 8, txt=str(uvalu), align="J")
        pdf.set_font("Arial", size=16)
        filename=str(ind)+" "+str(avalu)+".pdf"
        pdf.output(filename.lower())
        pdfvar.set(filename.lower()+"       SAVED!!!")

frame1=Frame(window)

def excel():
    global tname,fname,pdfvar,excelvar,avalu,bvalu,cvalu,dvalu,evalu,fvalu,gvalu,hvalu,ivalu,jvalu,kvalu,lvalu,mvalu,nvalu,ovalu,pvalu,qvalu,rvalu,svalu,tvalu,uvalu,vvalu,wvalu
    avalu=a1.get()
    bvalu=b1.get()
    cvalu=c1.get()
    dvalu=d1.get()
    evalu=e1.get()
    fvalu=f1.get()
    gvalu=g1.get()
    hvalu=h1.get()
    ivalu=variable.get()
    if ivalu =="This is default Option Do not choose this.":
        ivalu=None
    jvalu=variable2.get()
    if jvalu =="This is default Option Do not choose this.":
        jvalu=None
    kvalu=k1.get()
    lvalu=l1.get()
    mvalu=m1.get()
    nvalu=n1.get()
    ovalu=o1.get()
    pvalu=p1.get()
    qvalu=q1.get()
    rvalu=r1.get()
    svalu=s1.get()
    tvalu=t1.get()
    uvalu=u1.get(1.0,"end")
    tmp=ovalu.split(".")
    if len(tmp)==3 and len(tmp[2])==2:
        tmp[2]="20"+tmp[2]
        a=ovalu.split(".")
        a[2]=" 20"+a[2]
        ovalu=".".join(a)
        tmp=tmp[::-1]
        tmp="-".join(tmp)
        df=pd.Timestamp(tmp)
        vvalu=str(df.day_name())
    wvalu=now.strftime("%d/%m/%Y %H:%M:%S %p %a")
    #print(avalu,bvalu,cvalu,dvalu,evalu,fvalu,gvalu,hvalu,ivalu,jvalu,kvalu,lvalu,mvalu,nvalu,ovalu,pvalu,qvalu,rvalu,svalu,tvalu,uvalu,vvalu,wvalu)
    fname="registry_list.xls"
    tname="registry_list.xlsx"
    workbook = xlsxwriter.Workbook(fname)    
    if os.path.isfile(fname) or os.path.isfile(tname) :
        pass
    else:
        workbook2=xlwt.Workbook(fname)
        ws = workbook2.add_sheet('Tested')
        ws.write(0,0,"S.No.")
        ws.write(0,1,"First Party Name")
        ws.write(0,2,"First Party Mobile Number")
        ws.write(0,3,"First Party Aadhar Number")
        ws.write(0,4,"First Party Pan Card")
        ws.write(0,5,"Second Party Name")
        ws.write(0,6,"Second Party Number")
        ws.write(0,7,"Second Party Aadhar Number")
        ws.write(0,8,"Second Party Pan Card")
        ws.write(0,9,"Document Date")
        ws.write(0,10,"Registry Day")
        ws.write(0,11,"Document")
        ws.write(0,12,"Deed Writer Number")
        ws.write(0,13,"Sub Registrar")
        ws.write(0,14,"Registration Number")
        ws.write(0,15,"Property Value")
        ws.write(0,16,"Maarfat")
        ws.write(0,17,"Entry Date")
        ws.write(0,18,"Property Address")
        ws.write(0,19,"Witness 1 Name")
        ws.write(0,20,"Witness 1 Number")
        ws.write(0,21,"Witness 2 Name")
        ws.write(0,22,"Witness 2 Number")
        workbook2.save(fname)
    if not os.path.isfile(tname):
        pd.read_excel(fname).to_excel(tname, sheet_name="Tested",index=False)
        os.remove(fname)
    if len(tvalu)==0:
        dataframe=pd.read_excel(tname)
        dataframe=dataframe['S.No.'].tolist()
        if len(dataframe)==0:
            ind=1
        else:
            dataframe=[int(x) for x in dataframe if not np.isnan(x)]
            dataframe=max(dataframe)
            ind=dataframe+1
    else:
        ind=tvalu
    indind=int(ind)+1
    if len(avalu)!=0:
        wb = load_workbook(tname)
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[0]]
        Sheet1 .cell(row = indind, column = 1).value = ind
        Sheet1 .cell(row = indind, column = 2).value = avalu
        Sheet1 .cell(row = indind, column = 3).value = bvalu
        Sheet1 .cell(row = indind, column = 4).value = cvalu
        Sheet1 .cell(row = indind, column = 5).value = dvalu
        Sheet1 .cell(row = indind, column = 6).value = evalu
        Sheet1 .cell(row = indind, column = 7).value = fvalu
        Sheet1 .cell(row = indind, column = 8).value = gvalu
        Sheet1 .cell(row = indind, column = 9).value = hvalu
        Sheet1 .cell(row = indind, column = 10).value = ovalu
        Sheet1 .cell(row = indind, column = 11).value = vvalu
        Sheet1 .cell(row = indind, column = 12).value = ivalu
        Sheet1 .cell(row = indind, column = 13).value = pvalu
        Sheet1 .cell(row = indind, column = 14).value = jvalu
        Sheet1 .cell(row = indind, column = 15).value = qvalu
        Sheet1 .cell(row = indind, column = 16).value = rvalu
        Sheet1 .cell(row = indind, column = 17).value = svalu
        Sheet1 .cell(row = indind, column = 18).value = wvalu
        Sheet1 .cell(row = indind, column = 19).value = uvalu
        Sheet1 .cell(row = indind, column = 20).value = kvalu
        Sheet1 .cell(row = indind, column = 21).value = lvalu
        Sheet1 .cell(row = indind, column = 22).value = mvalu
        Sheet1 .cell(row = indind, column = 23).value = nvalu
        wb.save(tname)
        excelvar.set("SAVED!!!")
       



    

def reset():
    global listy , listy2, variable, variable2,excelvar,pdfvar
    variable.set(listy[0])
    variable2.set(listy2[0])
    a1.delete(0,"end")
    b1.delete(0,"end")
    c1.delete(0,"end")
    d1.delete(0,"end")
    e1.delete(0,"end")
    f1.delete(0,"end")
    g1.delete(0,"end")
    h1.delete(0,"end")
    k1.delete(0,"end")
    l1.delete(0,"end")
    m1.delete(0,"end")
    n1.delete(0,"end")
    o1.delete(0,"end")
    p1.delete(0,"end")
    q1.delete(0,"end")
    r1.delete(0,"end")
    s1.delete(0,"end")
    t1.delete(0,"end")
    u1.delete(1.0,"end")
    excelvar.set("Do You Want To Save Excel File?")
    pdfvar.set("Do You Want To Save PDF File?")
    


listy = ["This is default Option Do not choose this.","Adoption","Affidavit","Agreement","Sale Deed (Conveyance Deed)","Sale Deed (Female SC/ST/BPL)","Sale Deed (Female other than SC/ST/BPL)","Sale Deed (Disable 40% & Above)","Exchange Deed","Gift Deed","Lease deed below 1 years","Lease deed 1 years to 5 years","Lease deed exceeding 5 years to 10 years","Lease deed exceeding 10 years to 15 years","Lease deed exceeding 15 years to 20 years","Lease deed exceed 20 years to 30 years","Lease deed exceeding 30 years and perpetual","Mortgage Deed","Partnership","Power of Attorney","Release Deed","Will"]
variable = StringVar(window)
variable.set(listy[0])

listy2= ["This is default Option Do not choose this.","AMER","BAGRU KALAN","BASSI","CHAKSU","CHOMU","CHTOH KA BARWADA","DUDU","GOVINDGARH","JAIPUR I","JAIPUR II","JAIPUR III","JAIPUR IV","JAIPUR V","JAIPUR VI","JAIPUR VII","JAIPUR VIII","JAIPUR IX","JAIPUR X","SANGANER I","SANGANER II","JALSU","JAMWARAM GARH","KALWAR","KISHAN GARH RENWAL","KOTKHAWAD","KOTPUTALI","MADHORAJPURA","MOJMABAD","PAWTA","PHAGI","RAMPURA DABRI","SHAHPURA","SAMBHAR","TUNGA","VIRATNAGAR","MALPURA"]
variable2 = StringVar(window)
variable2.set(listy[0])

def defocus(event):
    event.widget.master.focus_set()

a = Label(frame1 ,text = "First Party Name",height=2,width=26).grid(row = 0,column = 0)
b = Label(frame1 ,text = "First Party Mobile Number",height=2,width=26).grid(row = 1,column = 0)
c = Label(frame1 ,text = "First Party Aadhar Number",height=2,width=26).grid(row = 2,column = 0)
d = Label(frame1 ,text = "First Party Pan Card",height=2,width=26).grid(row = 3,column = 0)
e = Label(frame1 ,text = "Second Party Name",height=2,width=26).grid(row = 4,column = 0)
f = Label(frame1 ,text = "Second Party Mobile Number",height=2,width=26).grid(row = 5,column = 0)
g = Label(frame1 ,text = "Second Party Aadhar Number",height=2,width=26).grid(row = 6,column = 0)
h = Label(frame1 ,text = "Second Party Pan Card",height=2,width=26).grid(row = 7,column = 0)
i = Label(frame1 ,text = "Document",height=2,width=26).grid(row = 8,column = 0)
j = Label(frame1 ,text = "Sub Registrar",height=2,width=26).grid(row = 9,column = 0)

k = Label(frame1 ,text = "Witness 1 Name",height=2,width=26).grid(row = 0,column = 3)
l = Label(frame1 ,text = "Witness 1 Mobile Number",height=2,width=26).grid(row = 1,column = 3)
m = Label(frame1 ,text = "Witness 2 Name",height=2,width=26).grid(row = 2,column = 3)
n = Label(frame1 ,text = "Witness 2 Mobile Number",height=2,width=26).grid(row = 3,column = 3)
o = Label(frame1 ,text = "Document Date ( . )",height=2,width=26).grid(row = 4,column = 3)
p = Label(frame1 ,text = "Deed Writer Number",height=2,width=26).grid(row = 5,column = 3)
q = Label(frame1 ,text = "Registration Number",height=2,width=26).grid(row = 6,column = 3)
r = Label(frame1 ,text = "Property Value",height=2,width=26).grid(row = 7,column = 3)
s = Label(frame1 ,text = "Maarfat",height=2,width=26).grid(row = 8,column = 3)
t = Label(frame1 ,text = "S.No.",height=2,width=26).grid(row = 9,column = 3)
u = Label(frame1 ,text = "Property Address",height=2,width=26).grid(row =10,column = 3)



a1 = Entry(frame1,width=45,highlightthickness=7)
a1.grid(row = 0,column = 1)
b1 = Entry(frame1,width=45,highlightthickness=7)
b1.grid(row = 1,column = 1)
c1 = Entry(frame1,width=45,highlightthickness=7)
c1.grid(row = 2,column = 1)
d1 = Entry(frame1,width=45,highlightthickness=7)
d1.grid(row = 3,column = 1)
e1 = Entry(frame1,width=45,highlightthickness=7)
e1.grid(row = 4,column = 1)
f1 = Entry(frame1,width=45,highlightthickness=7)
f1.grid(row = 5,column = 1)
g1 = Entry(frame1,width=45,highlightthickness=7)
g1.grid(row = 6,column = 1)
h1 = Entry(frame1,width=45,highlightthickness=7)
h1.grid(row = 7,column = 1)
i1 = ttk.Combobox(frame1,textvariable=variable,values=listy)
i1.config(width=43)
#i1.bind("<FocusIn>", defocus)
i1.grid(row = 8,column = 1)
j1 = ttk.Combobox(frame1,textvariable=variable2,values=listy2)
j1.config(width=43)
#j1.bind("<FocusIn>", defocus)
j1.grid(row = 9,column = 1)

k1 = Entry(frame1,width=45,highlightthickness=7)
k1.grid(row = 0,column = 4)
l1 = Entry(frame1,width=45,highlightthickness=7)
l1.grid(row = 1,column = 4)
m1 = Entry(frame1,width=45,highlightthickness=7)
m1.grid(row = 2,column = 4)
n1 = Entry(frame1,width=45,highlightthickness=7)
n1.grid(row = 3,column = 4)
o1 = Entry(frame1,width=45,highlightthickness=7)
o1.grid(row = 4,column = 4, padx=10)
p1 = Entry(frame1,width=45,highlightthickness=7)
p1.grid(row = 5,column = 4)
q1 = Entry(frame1,width=45,highlightthickness=7)
q1.grid(row = 6,column = 4)
r1 = Entry(frame1,width=45,highlightthickness=7)
r1.grid(row = 7,column = 4)
s1 = Entry(frame1,width=45,highlightthickness=7)
s1.grid(row = 8,column = 4)
t1 = Entry(frame1,width=45,highlightthickness=7)
t1.grid(row = 9,column = 4)
u1 = Text(frame1,width=45,height=4,highlightthickness=7,)
u1.grid(row = 10,column = 4,padx=10)

frame1.pack(expand=True,side="top",fill="both")







frame2=Frame(window)


excel=Button(frame2,text="EXCEL FILE",width=40,highlightthickness=7,command=excel).grid(row = 0,column = 0,padx=30)
reset=Button(frame2,text="RESET BUTTON",width=40,highlightthickness=7,command=reset).grid(row = 0,column = 1,padx=30)
pdf=Button(frame2,text="PDF FILE",width=40,highlightthickness=7,command=pdf).grid(row = 0,column = 2,padx=30)

excelvar=StringVar()
pdfvar=StringVar()
excelvar.set("Do You Want To Save Excel File?")
pdfvar.set("Do You Want To Save PDF File?")
excel1=Label(frame2 ,textvariable=excelvar,height=2,width=40 )
excel1.grid(row = 1,column = 0)
pdf1=Label(frame2 ,textvariable=pdfvar,height=2,width=60)
pdf1.grid(row = 1,column = 2)


frame2.pack(expand=True,side="bottom",fill="both")



window.mainloop()
exceling = win32.gencache.EnsureDispatch('Excel.Application')
r=os.path.abspath('registry_list.xlsx')
wbook = exceling.Workbooks.Open(r)
wsheet = wbook.Worksheets("Tested")
wsheet.Columns.AutoFit()
wbook.Save()
exceling.Application.Quit()


